import sys
import re
from pathlib import Path
from datetime import datetime, date
import os

from openpyxl import load_workbook

import firebase_admin
from firebase_admin import credentials, firestore

# ==============================
# Firebase init
# ==============================
if not firebase_admin._apps:
    cred = credentials.Certificate({
        "type": "service_account",
        "project_id": os.environ["FIREBASE_PROJECT_ID"],
        "client_email": os.environ["FIREBASE_CLIENT_EMAIL"],
        "private_key": os.environ["FIREBASE_PRIVATE_KEY"].replace("\\n", "\n"),
    })
    firebase_admin.initialize_app(cred)

db = firestore.client()

# ==============================
# Required headers (table headers)
# ==============================
REQUIRED_HEADERS = {
    "staff": ["שם המרצה", "מרצה"],
    "group": ["קבוצת מעבדה", "קבוצה"],
    "time": ["שעה"],
    "day": ["יום"],
    "date": ["תאריך"],
    "sessionNo": ["מס' מע'", "מספר מעבדה", "מס׳ מע", "מס' מעבדה"],
    "sessionName": ["שם המקצוע", "שם הקורס"],  # אצלך זה "BSL2 הדרכה" / "תרביות תאים" וכו'
}

# ==============================
# Helpers
# ==============================
def norm(x) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()

def parse_date_to_iso(v) -> str:
    """Return YYYY-MM-DD if possible, else original normalized string."""
    if isinstance(v, (datetime, date)):
        # openpyxl may give datetime/date
        d = v.date() if isinstance(v, datetime) else v
        return d.isoformat()

    s = norm(v)
    if not s:
        return ""

    # Supports dd.mm.yy or dd.mm.yyyy
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", s)
    if m:
        dd, mm, yy = m.groups()
        dd = int(dd); mm = int(mm); yy = int(yy)
        if yy < 100:
            yy += 2000  # 25 -> 2025
        try:
            return date(yy, mm, dd).isoformat()
        except Exception:
            return s

    return s

def build_header(ws, row, max_col):
    cells = [norm(ws.cell(row=row, column=c).value) for c in range(1, max_col + 1)]
    mapping = {}
    hits = 0

    for key, variants in REQUIRED_HEADERS.items():
        for i, txt in enumerate(cells, start=1):
            if not txt:
                continue
            if any(v in txt for v in variants):
                mapping[key] = i
                hits += 1
                break

    return mapping, hits

def is_table_header_row(hits: int) -> bool:
    # 5/7 headers is usually enough to detect a table start
    return hits >= 5

def extract_course_from_line(text: str):
    """
    Detect a course title line like:
      "תרביות תאים - 41262"
      "41262 - תרביות תאים"
    Returns (course_code, course_name) or (None, None)
    """
    t = norm(text)
    if not t:
        return None, None

    # try: name - code
    m = re.search(r"(.+?)\s*[-–]\s*(\d{4,6})$", t)
    if m:
        name = norm(m.group(1))
        code = norm(m.group(2))
        return code, name

    # try: code - name
    m = re.search(r"^(\d{4,6})\s*[-–]\s*(.+)$", t)
    if m:
        code = norm(m.group(1))
        name = norm(m.group(2))
        return code, name

    return None, None

def find_course_title_near(ws, header_row: int, max_col: int):
    """
    Look up to ~8 rows above the table header for a line containing course name + code.
    Excel files often have the title above the table (not inside it).
    """
    for r in range(header_row - 1, max(1, header_row - 8), -1):
        row_texts = []
        for c in range(1, max_col + 1):
            v = norm(ws.cell(row=r, column=c).value)
            if v:
                row_texts.append(v)

        # Sometimes the title is in one cell, sometimes in several.
        line = " ".join(row_texts)
        code, name = extract_course_from_line(line)
        if code and name:
            return code, name

    return None, None

# ==============================
# Core logic (OLD STRUCTURE)
# ==============================
def parse_workbook(path: Path, year_id: str, year_label: str, semester: str):
    wb = load_workbook(path, data_only=True)

    # Root year doc
    year_ref = db.collection("lab_schedule").document(year_id)
    year_ref.set({"year": year_label}, merge=True)

    semester_doc_ref = year_ref.collection("semesters").document(str(semester))

    # We'll build ONE doc with courses map
    courses_map = {}  # { "41262": {courseCode, courseName, labs:[...] } }

    for ws in wb.worksheets:
        r = 1
        while r <= ws.max_row:
            header, hits = build_header(ws, r, ws.max_column)

            if is_table_header_row(hits):
                # Find course code+name above this header
                course_code, course_name = find_course_title_near(ws, r, ws.max_column)

                # If we couldn't find a title, we still can continue,
                # but we need a key. We'll skip to avoid broken structure.
                if not course_code or not course_name:
                    # move on
                    r += 1
                    continue

                # Ensure course entry exists
                if course_code not in courses_map:
                    courses_map[course_code] = {
                        "courseCode": str(course_code),
                        "courseName": course_name,
                        "labs": [],
                    }

                rr = r + 1
                while rr <= ws.max_row:
                    # Stop when "date" and "day" are empty and also sessionName empty => end of table block
                    session_name = norm(ws.cell(rr, header.get("sessionName", 0)).value) if header.get("sessionName") else ""
                    dval = ws.cell(rr, header.get("date", 0)).value if header.get("date") else None
                    dayval = norm(ws.cell(rr, header.get("day", 0)).value) if header.get("day") else ""

                    if not session_name and not norm(dval) and not dayval:
                        break

                    # Build lab row
                    lab = {
                        "date": parse_date_to_iso(dval),
                        "day": dayval,
                        "group": norm(ws.cell(rr, header.get("group", 0)).value) if header.get("group") else "",
                        "time": norm(ws.cell(rr, header.get("time", 0)).value) if header.get("time") else "",
                        "staff": [],
                        # session: in your old DB you used "session" as the thing in table (BSL2/תרביות תאים)
                        "session": session_name,
                    }

                    # session number (optional)
                    if header.get("sessionNo"):
                        lab["sessionNo"] = norm(ws.cell(rr, header["sessionNo"]).value)

                    # staff (optional)
                    if header.get("staff"):
                        staff = norm(ws.cell(rr, header["staff"]).value)
                        if staff:
                            lab["staff"] = [staff]

                    # Clean tiny empties
                    if not lab["date"]:
                        lab["date"] = ""

                    courses_map[course_code]["labs"].append(lab)
                    rr += 1

                r = rr  # continue after this block
            else:
                r += 1

    # Save EXACT old structure: one semester doc with courses map
    semester_doc_ref.set(
        {
            "semester": int(semester),
            "updatedAt": firestore.SERVER_TIMESTAMP,
            "courses": courses_map,
        },
        merge=False,  # replace doc completely (like old behavior)
    )

# ==============================
# ENTRY POINT
# ==============================
if __name__ == "__main__":
    if len(sys.argv) < 5:
        raise Exception("Usage: labs_parser.py <file_path> <year_id> <year_label> <semester>")

    file_path = sys.argv[1]
    year_id = sys.argv[2]
    year_label = sys.argv[3]
    semester = sys.argv[4]

    parse_workbook(Path(file_path), year_id, year_label, semester)
    print("OK")
